import time
import cv2
from ultralytics import YOLO
import numpy as np
from PIL import ImageFont, ImageDraw, Image
import os
import re
import logging
from openpyxl import Workbook, load_workbook

logging.getLogger("ultralytics").setLevel(logging.WARNING)

find_plate_model = YOLO("./model/best_plate.pt")
plate_char_model = YOLO("./model/best_char.pt")

classes_type = ['plate1', 'plate2', 'plate3']
classes_char = ['1','2','3','4','5','6','7','8','9','0',
           '가', '나', '다', '라', '마',
           '거', '너', '더', '러', '머', '버', '서', '어', '저',
           '고', '노', '도', '로', '모', '보', '소', '오', '조',
           '구', '누', '두', '루', '무', '부', '수', '우', '주',
           '아', '바', '사', '자', '허', '하', '호', '배', 
           '서울', '서울', '부산', '부산', '대구', '대구', '인천', '인천', 
           '광주', '광주', '대전', '대전', '울산', '울산', '세종', '세종', 
           '경기', '경기', '강원', '강원', '충북', '충북', '충남', '충남', 
           '전북', '전북', '전남', '전남', '경북', '경북', '경남','경남', 
           '제주', '제주', '육', '해', '공', '국', '합', '초', '퍼', 
           '처', '으', '토', '흐', '느', '후', '타', '파', '차', '커', '크', '스', '코']

COLOR = (255, 255, 255, 0)
FONT_SIZE = 70
MISMATCH_COLOR = (0, 0, 255)

roi_pts = []
roi_set = False
scale_factor = 1.0

def click_event(event, x, y, flags, param):
    global roi_pts, roi_set
    if event == cv2.EVENT_LBUTTONDOWN:
        roi_pts.append((int(x / scale_factor), int(y / scale_factor)))
        if len(roi_pts) >= 4:
            roi_set = True

def detect_char_v8(frame, plate_type):
    obj_list = []
    results = plate_char_model(frame)

    for result in results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5])
            if conf >= 0.6:
                label = classes_char[cls]
                obj_list.append([label, x1, y1, x2, y2])

    if obj_list:
        if plate_type == 0:
            sorted_by_x = sorted(obj_list, key=lambda x: x[1])
            plate_number = ''.join([i[0] for i in sorted_by_x]).replace(" ", "")
        elif plate_type == 1:
            top_line = [item for item in obj_list if item[2] < (max(item[2] for item in obj_list) + min(item[2] for item in obj_list)) / 2]
            bottom_line = [item for item in obj_list if item[2] >= (max(item[2] for item in obj_list) + min(item[2] for item in obj_list)) / 2]
            top_line.sort(key=lambda x: x[1])
            bottom_line.sort(key=lambda x: x[1])
            plate_number = ''.join([i[0] for i in top_line + bottom_line]).replace(" ", "")
        elif plate_type == 2:
            sorted_by_x = sorted(obj_list, key=lambda x: x[1])
            plate_number = ''.join([i[0] for i in sorted_by_x]).replace(" ", "")
        return plate_number
    else:
        logging.info("list empty!")
        return None

def is_within_roi(bbox, roi_pts, img):
    if not roi_pts:
        return True
    x1, y1, x2, y2 = bbox
    roi_polygon = np.array(roi_pts, dtype=np.int32)
    box_polygon = np.array([(x1, y1), (x2, y1), (x2, y2), (x1, y2)], dtype=np.int32)
    box_mask = np.zeros((img.shape[0], img.shape[1]), dtype=np.uint8)
    roi_mask = np.zeros((img.shape[0], img.shape[1]), dtype=np.uint8)
    cv2.fillPoly(box_mask, [box_polygon], 1)
    cv2.fillPoly(roi_mask, [roi_polygon], 1)
    intersection = np.logical_and(box_mask, roi_mask)
    return np.any(intersection)

def detect_plate_and_keypoints(img_path):
    global roi_pts, roi_set, scale_factor
    
    now_AI = time.strftime('%Y-%m-%d %H:%M:%S')
    frame = cv2.imread(img_path, cv2.IMREAD_COLOR)
    img_width = frame.shape[1]
    
    results = find_plate_model(frame)
    
    if len(results[0].boxes) == 0:
        logging.info("[detect_plate_and_keypoints] No plates detected.")
        return None, None, img_width
    
    position_list = []
    obj_list = []
    crop_images = []

    for i, box in enumerate(results[0].boxes.data):
        x1, y1, x2, y2, conf, cls = int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5])
        if conf >= 0.6:
            label = classes_type[cls]
            
            if label == "plate1":
                plate_type = 0
            elif label == "plate2":
                plate_type = 1
            elif label == "plate3":
                plate_type = 2
                
            if i < len(results[0].keypoints.data):
                keypoint = results[0].keypoints.data[i]
                kx1, ky1, kx2, ky2, kx3, ky3, kx4, ky4 = int(keypoint[0][0]), int(keypoint[0][1]), int(keypoint[1][0]), int(keypoint[1][1]), int(keypoint[2][0]), int(keypoint[2][1]), int(keypoint[3][0]), int(keypoint[3][1])
                points_value = [kx1, ky1], [kx2, ky2], [kx3, ky3], [kx4, ky4]

                if [0, 0] in points_value:
                    crop = frame[y1:y2, x1:x2]
                    if is_within_roi((x1, y1, x2, y2), roi_pts, frame):
                        plate_number = detect_char_v8(crop, plate_type)
                        position_list.append([label, plate_number, x1, y1, x2, y2])
                        crop_images.append(crop)
                        obj_list.append([label, plate_number, x1, y1, x2, y2, plate_type])
                else:
                    points_src = np.array([[kx1, ky1], [kx3, ky3], [kx2, ky2], [kx4, ky4]], dtype=np.float32)
                    w1, w2 = abs(kx4 - kx2), abs(kx3 - kx1)
                    h1, h2 = abs(ky3 - ky4), abs(ky1 - ky2)
                    height, width = max([h1, h2]), max([w1, w2])
                    pts2 = np.float32([[0, 0], [width - 1, 0], [0, height - 1], [width - 1, height - 1]])
                    matrix = cv2.getPerspectiveTransform(points_src, pts2)
                    crop_img = cv2.warpPerspective(frame, matrix, (width, height))
                    if is_within_roi((x1, y1, x2, y2), roi_pts, frame):
                        plate_number = detect_char_v8(crop_img, plate_type)
                        crop_images.append(crop_img)
                        obj_list.append([label, plate_number, x1, y1, x2, y2, plate_type])
    
    if crop_images and obj_list:
        return crop_images, obj_list, img_width
    else:
        logging.info("[detect_plate_and_keypoints] No plates detected.")
        return None, None, img_width

def position(position_list, img_width):
    if len(position_list) == 1:
        center_x = img_width / 2
        obj_x1 = position_list[0][2]
        position = "left" if obj_x1 < center_x else "right"
        #logging.info(f"검출 차량 번호: {position_list[0][1]}, 위치: {position}")

    elif len(position_list) == 2:
        sorted_positions = sorted(position_list, key=lambda x: x[2])
        for position_info in position_list:
            x1 = position_info[2]
            position = "left" if x1 == sorted_positions[0][2] else "right"
            #logging.info(f"검출 차량 번호: {position_info[1]}, 위치: {position}")

    else:
        sorted_positions = sorted(position_list, key=lambda x: x[2])
        for position_info in position_list:
            x1 = position_info[2]
            if x1 == sorted_positions[0][2]:
                position = "left"
            elif x1 == sorted_positions[-1][2]:
                position = "right"
            else:
                position = "center"
            #logging.info(f"검출 차량 번호: {position_info[1]}, 위치: {position}")

def myPutText(src, text, pos, font_size, font_color):
    if text is None:
        text = "N/A"
    img_pil = Image.fromarray(src)
    draw = ImageDraw.Draw(img_pil)
    font = ImageFont.truetype('fonts/gulim.ttc', font_size)
    bbox = draw.textbbox(pos, text, font=font)
    draw.rectangle(bbox, fill="red")
    draw.text(pos, text, font=font, fill=font_color)
    return np.array(img_pil)

def save_crop(img, original_file, suffix):
    base_name = os.path.splitext(os.path.basename(original_file))[0]
    crop_dir = os.path.join(os.path.dirname(original_file), "crop")
    os.makedirs(crop_dir, exist_ok=True)
    
    count = 1
    crop_name = f"{base_name}_{suffix}_{count}.jpg"
    while os.path.exists(os.path.join(crop_dir, crop_name)):
        count += 1
        crop_name = f"{base_name}_{suffix}_{count}.jpg"
    
    cv2.imwrite(os.path.join(crop_dir, crop_name), img)


def extract_gt_from_filename(filename, electric_plates):
    base_name = os.path.splitext(os.path.basename(filename))[0]
    parts = base_name.split('_')
    plate_pattern = re.compile(r'\d{2,3}[가-힣]\d{4}')
    gt_list = [part for part in parts if plate_pattern.match(part)]
    return [(plate, '전기차량' if plate in electric_plates else '불법주차차량') for plate in gt_list]

def compare_results_with_gt(result_list, gt_list):
    tp_illegal = 0
    fp_illegal = 0
    fn_illegal = 0

    tp_electric = 0
    fp_electric = 0
    fn_electric = 0
    
    gt_illegal = [plate for plate in gt_list if plate[1] == '불법주차차량']
    gt_electric = [plate for plate in gt_list if plate[1] == '전기차량']
    
    detected_illegal = [plate for plate in result_list if plate[1] == '불법주차차량']
    detected_electric = [plate for plate in result_list if plate[1] == '전기차량']

    if len(gt_list) > len(result_list):
        for gt_plate in gt_list:
            if gt_plate not in result_list:
                if gt_plate[1] == '불법주차차량':
                    fn_illegal += 1
                elif gt_plate[1] == '전기차량':
                    fn_electric += 1
    
    detected_pass = []
    for detected in detected_illegal:
        if detected in gt_illegal:
            tp_illegal += 1
            gt_illegal.remove(detected)
        else:
            detected_pass.append(detected)

    for detected in detected_electric:
        if detected in gt_electric:
            tp_electric += 1
            gt_electric.remove(detected)
        else:
            detected_pass.append(detected)

    for detected in detected_pass:
        if gt_electric:
            fp_electric += 1
            gt_electric.pop(0)
        elif gt_illegal:
            fp_illegal += 1
            gt_illegal.pop(0)

    return tp_illegal, fp_illegal, fn_illegal, tp_electric, fp_electric, fn_electric
   
def result_and_save_image(img, position_list, save_path, accuracy_illegal=None, accuracy_electric=None, electric_plates=[]):
    plate_number_list = []
    result_details = []
    
    for pos_info in position_list:
        plate_number = pos_info[1]
        plate_type = pos_info[0]
        plate_number_list.append(plate_number)
        img = myPutText(img, plate_number, (pos_info[2], pos_info[3] - 60), FONT_SIZE, COLOR)
        img = cv2.rectangle(img, (pos_info[2], pos_info[3]), (pos_info[4], pos_info[5]), (255, 0, 0), 3)
            
        if plate_type == 'plate3':
            img = myPutText(img, "전기 차량", (pos_info[2], pos_info[5] + 10), FONT_SIZE, (0, 0, 255))
            result_details.append(f"{plate_number}:전기차량")
        else:
            img = myPutText(img, "불법 주차 차량", (pos_info[2], pos_info[5] + 10), FONT_SIZE, (0, 0, 255))
            result_details.append(f"{plate_number}:불법주차차량")
    
    gt_text = "GT: " + ", ".join([f"{plate[0]}:{plate[1]}" for plate in gt_list])
    result_text_parts = []
    correct_matches = 0    
    
    for res in result_details:
        if res in gt_text:
            result_text_parts.append((res, COLOR))
            correct_matches += 1
        else:
            result_text_parts.append((res, MISMATCH_COLOR))

    gt_color = COLOR
    result_text = "Result: "
    for part in result_text_parts:
        result_text += f"{part[0]}, "

    img = myPutText(img, gt_text, (10, 10), FONT_SIZE, gt_color)
    
    y_position = 70
    img = myPutText(img, "Result: ", (10, y_position), FONT_SIZE, COLOR)
    
    x_offset = 256
    for part in result_text_parts:
        img = myPutText(img, part[0], (x_offset, y_position), FONT_SIZE, part[1])
        x_offset += 50 * (len(part[0]) + 1)
    
    if roi_pts:
        img = cv2.polylines(img, [np.array(roi_pts, dtype=np.int32)], isClosed=True, color=(0, 255, 0), thickness=2)
    
    cv2.imwrite(save_path, img)
    
    img_name = os.path.basename(save_path)
    gt = [f"{plate[0]}:{plate[1]}" for plate in gt_list]
    result_summary = ", ".join(result_details)
    gt_summary = ", ".join(gt)
    total_gt = len(gt_list)
    average_accuracy = correct_matches / total_gt if total_gt > 0 else 0
    
    excel_path = os.path.join(os.path.dirname(save_path), "plate_result.xlsx")
    
    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Image Name", "Ground Truth", "Recognized Plate Number", "Accuracy"])
    
    sheet.append([img_name, gt_summary, result_summary, average_accuracy])    
    workbook.save(excel_path)



def state(img_file, result_folder, electric_plates):
    global roi_pts, roi_set, scale_factor
    global gt_list

    img = cv2.imread(img_file)
    
    if img is not None:
        img_name = os.path.basename(img_file)
        logging.info('')
        logging.info(f'<image: {img_name}>')
        
        save_path = os.path.join(result_folder, img_name)
        
        crop_images, obj_list, img_width = detect_plate_and_keypoints(img_file)
        gt_list = extract_gt_from_filename(img_file, electric_plates)
        
        detected_plate_numbers = []

        for result in (obj_list or []):
            plate_number = result[1]
            plate_type = result[0]
            
            if plate_type == 'plate3':
                plate_category = '전기차량'
            else:
                plate_category = '불법주차차량'
            
            detected_plate_numbers.append((plate_number, plate_category))

        logging.info(f"Detected Plates: {detected_plate_numbers}")
        logging.info(f"Ground Truth Plates: {gt_list}")

        if crop_images and obj_list:
            tp_illegal, fp_illegal, fn_illegal, tp_electric, fp_electric, fn_electric = compare_results_with_gt(detected_plate_numbers, gt_list)
            accuracy_illegal = tp_illegal / (tp_illegal + fn_illegal + fp_illegal) if (tp_illegal + fn_illegal + fp_illegal) else 0
            accuracy_electric = tp_electric / (tp_electric + fn_electric + fp_electric) if (tp_electric + fn_electric + fp_electric) else 0
            
            logging.info(f"Illegal Plates - (TP: {tp_illegal}, FP: {fp_illegal}, FN: {fn_illegal})")
            logging.info(f"Electric Plates - (TP: {tp_electric}, FP: {fp_electric}, FN: {fn_electric})")
            result_and_save_image(img, obj_list, save_path, accuracy_illegal, accuracy_electric, electric_plates)
            position(obj_list, img_width)
        else:
            fn_illegal = sum(1 for gt_plate in gt_list if gt_plate[1] == '불법주차차량')
            fn_electric = sum(1 for gt_plate in gt_list if gt_plate[1] == '전기차량')
            tp_illegal, fp_illegal, tp_electric, fp_electric = 0, 0, 0, 0
            logging.info(f"Illegal Plates - (TP: 0, FP: 0, FN: {fn_illegal})")
            logging.info(f"Electric Plates - (TP: 0, FP: 0, FN: {fn_electric})")
            logging.info("GT 값은 있지만 검출된 번호판이 없습니다.")
            result_and_save_image(img, [], save_path, 0, 0, electric_plates)
        
        return tp_illegal, fp_illegal, fn_illegal, tp_electric, fp_electric, fn_electric

    else:
        logging.info(f"Failed to load image: {img_file}")
        return 0, 0, 0, 0, 0, 0
    
def save_overall_results(test_folder):
    """
    Save the overall results to the test folder.
    """
    overall_excel_path = os.path.join(test_folder, "plate_result.xlsx")
    if os.path.exists(overall_excel_path):
        workbook = load_workbook(overall_excel_path)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Dataset", "Image Name", "Ground Truth", "Recognized Plate Number", "Accuracy"])
    
    for dataset_folder in os.listdir(test_folder):
        dataset_path = os.path.join(test_folder, dataset_folder)
        if os.path.isdir(dataset_path):
            excel_path = os.path.join(dataset_path, "plate_result.xlsx")
            if os.path.exists(excel_path):
                dataset_workbook = load_workbook(excel_path)
                dataset_sheet = dataset_workbook.active
                overall_sheet = workbook.active
                for row in dataset_sheet.iter_rows(min_row=2, values_only=True):
                    overall_sheet.append([dataset_folder] + list(row))
    
    workbook.save(overall_excel_path)